"""The workbook a run produces.

Five sheets, each answering one question:

``Summary``    did the run as a whole work, and what is the headline count?
``Results``    one row per operator — the sheet to filter and sort.
``Issues``     only the failures, grouped by root cause, so the work is a list.
``Surfaces``   only the operators a user cannot reach from all three surfaces.
``Environment`` what was run, against what, when — so a report can be re-run.

Written with openpyxl rather than pandas: the value here is the formatting —
frozen headers, autofilters, conditional colour, column widths that make the
`why` column readable — and a DataFrame dump would need all of it bolted back
on afterwards.
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .harness import (
    FAIL, OUTPUT_BUDGET_BYTES, PASS, SKIPPED, UNBUILT, OperatorOutcome,
    RunReport,
)



# A restrained palette: status colours have to be distinguishable at a glance
# without turning the sheet into a traffic light.
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SECTION_FONT = Font(bold=True, size=11, color="1F3864")

STATUS_FILLS = {
    PASS: PatternFill("solid", fgColor="C6EFCE"),
    FAIL: PatternFill("solid", fgColor="FFC7CE"),
    SKIPPED: PatternFill("solid", fgColor="FFEB9C"),
}
STATUS_FONTS = {
    PASS: Font(color="006100", bold=True),
    FAIL: Font(color="9C0006", bold=True),
    SKIPPED: Font(color="9C6500", bold=True),
}

YES_FONT = Font(color="006100")
NO_FONT = Font(color="9C0006", bold=True)
UNKNOWN_FONT = Font(color="808080", italic=True)

THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

#: ``(header, width, attribute-or-callable)`` for the Results sheet.
#: Widths are chosen so the sheet is readable without touching a column:
#: `Why` gets the room because it is the column the run exists to produce.
RESULT_COLUMNS: Sequence = (
    ("#", 5, None),
    ("Operator", 22, lambda row: row.operator),
    ("Status", 10, lambda row: row.status),
    ("Why", 62, lambda row: row.why),
    # Beside "Why" rather than at the end: it is the same kind of fact — what
    # went wrong — for a row whose status says nothing did. A pass with
    # something in this column is a pass worth reading.
    ("Pass-through", 62, lambda row: row.passthrough),
    ("Issue type", 30, lambda row: row.issue_type),
    ("Category", 20, lambda row: row.category),
    ("In toolbar", 11, lambda row: _tick(row.surfaces.toolbar)),
    ("Toolbar placement", 16, lambda row: row.surfaces.toolbar_placement),
    ("In model builder", 15, lambda row: _tick(row.surfaces.model_builder)),
    ("In command palette", 17, lambda row: _tick(row.surfaces.palette)),
    ("In installed CDO", 15, lambda row: _tick(row.surfaces.installed)),
    ("Signature", 10, lambda row: row.signature),
    ("Input ext", 10, lambda row: row.input_extension),
    ("Output ext", 11, lambda row: row.output_extension),
    ("Output kind", 17, lambda row: row.output_kind),
    ("Syntax", 38, lambda row: row.syntax),
    ("Parameters used", 26, lambda row: row.parameters),
    ("Input files", 26, lambda row: row.input_files),
    ("Output path", 40, lambda row: row.output_path),
    ("Output bytes", 13, lambda row: row.output_bytes),
    ("Exit code", 10, lambda row: row.returncode),
    ("Seconds", 9, lambda row: row.duration),
    ("Description", 45, lambda row: row.description),
    ("Command", 90, lambda row: row.command),
)


def _tick(value: Optional[bool]) -> str:
    """Three-valued, because "not inspected" is not "not present"."""
    if value is None:
        return "?"
    return "yes" if value else "NO"


#: Control characters openpyxl refuses to write. CDO output is arbitrary text —
#: progress carriage returns, occasionally raw bytes from a mangled file — and
#: one of them anywhere in 943 rows would abort the whole workbook.
_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

#: Excel's own hard limit is 32767 characters. Nothing useful is that long, and
#: a captured stderr line can be, so cells are clipped well below it.
MAX_CELL_CHARS = 2000


def _megabytes(count: int) -> str:
    """A byte count as MB, or GB once MB stops being readable."""
    if count >= 1024 ** 3:
        return f"{count / 1024 ** 3:.2f} GB"
    return f"{count / 1024 ** 2:.1f} MB"


def _safe(value: Any) -> Any:
    """A value Excel will actually accept."""
    if not isinstance(value, str):
        return value
    cleaned = _ILLEGAL.sub(" ", value)
    if len(cleaned) > MAX_CELL_CHARS:
        return cleaned[:MAX_CELL_CHARS] + " […]"
    return cleaned


def write_report(report: RunReport, path: Path) -> Path:
    """Write ``report`` to ``path`` and return the path actually written.

    A workbook already open in Excel cannot be replaced on Windows and is
    awkward on macOS, so a locked target falls back to a timestamped sibling
    rather than losing the run.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    _summary_sheet(workbook.create_sheet("Summary"), report)
    _results_sheet(workbook.create_sheet("Results"), report.outcomes)
    _issues_sheet(workbook.create_sheet("Issues"), report)
    _surfaces_sheet(workbook.create_sheet("Surfaces"), report)
    _environment_sheet(workbook.create_sheet("Environment"), report)

    path = Path(path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(path)
    except (OSError, PermissionError):
        stamped = path.with_name(
            f"{path.stem}_{datetime.now():%H%M%S}{path.suffix}")
        workbook.save(stamped)
        return stamped
    return path


# --------------------------------------------------------------------------
# Sheets
# --------------------------------------------------------------------------

def _summary_sheet(sheet: Worksheet, report: RunReport) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 78

    sheet["A1"] = "CDO operator test report"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"{report.started:%Y-%m-%d %H:%M:%S} — {report.duration:.0f}s"
    sheet["A2"].font = Font(italic=True, color="808080")

    passed, failed, skipped = (report.count(status) for status in (PASS, FAIL, SKIPPED))
    attempted = passed + failed
    rate = f"{passed / attempted * 100:.1f}%" if attempted else "n/a"

    # Two different kinds of skip, and they lead to different work: one is
    # fixed by rebuilding CDO with MAGICS or FFTW, the other by giving the
    # harness an input it does not know how to make. Counted apart so a reader
    # does not have to open the Issues sheet to tell which they are looking at.
    unbuilt = sum(1 for outcome in report.outcomes
                  if outcome.status == SKIPPED and outcome.issue_type == UNBUILT)

    row = _section(sheet, 4, "Headline")
    for label, value, status in (
        ("Operators in the run", len(report.outcomes), None),
        ("Passed", passed, PASS),
        ("Failed", failed, FAIL),
        ("Skipped (cannot be bulk-tested)", skipped - unbuilt, SKIPPED),
        ("Skipped (not in this CDO build)", unbuilt, SKIPPED),
        ("Pass rate of those attempted", rate, None),
        ("Unreachable from some surface", len(report.unreachable), None),
    ):
        sheet.cell(row, 1, label)
        cell = sheet.cell(row, 2, value)
        if status:
            cell.fill = STATUS_FILLS[status]
            cell.font = STATUS_FONTS[status]
        row += 1

    row = _section(sheet, row + 1, "Integration preflight")
    for check in report.preflight:
        sheet.cell(row, 1, check.name)
        cell = sheet.cell(row, 2, "OK" if check.ok else "PROBLEM")
        cell.fill = STATUS_FILLS[PASS if check.ok else FAIL]
        cell.font = STATUS_FONTS[PASS if check.ok else FAIL]
        sheet.cell(row, 3, _safe(check.detail)).alignment = Alignment(wrap_text=True)
        row += 1

    counts: dict = {}
    for outcome in report.failures:
        counts[outcome.issue_type or "Unclassified"] = counts.get(
            outcome.issue_type or "Unclassified", 0) + 1
    if counts:
        row = _section(sheet, row + 1, "Failures by root cause")
        for label, count in sorted(counts.items(), key=lambda item: -item[1]):
            sheet.cell(row, 1, label)
            sheet.cell(row, 2, count)
            row += 1


def _section(sheet: Worksheet, row: int, title: str) -> int:
    cell = sheet.cell(row, 1, title)
    cell.font = SECTION_FONT
    cell.border = Border(bottom=Side(style="medium", color="1F3864"))
    sheet.cell(row, 2).border = Border(bottom=Side(style="medium", color="1F3864"))
    return row + 1


def _results_sheet(sheet: Worksheet, outcomes: List[OperatorOutcome]) -> None:
    headers = [column[0] for column in RESULT_COLUMNS]
    _header_row(sheet, headers, [column[1] for column in RESULT_COLUMNS])

    for index, outcome in enumerate(outcomes, start=1):
        row = index + 1
        for column_index, (_, _, getter) in enumerate(RESULT_COLUMNS, start=1):
            value = index if getter is None else getter(outcome)
            cell = sheet.cell(row, column_index, _safe(value))
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top")

        status_cell = sheet.cell(row, headers.index("Status") + 1)
        status_cell.fill = STATUS_FILLS.get(outcome.status, STATUS_FILLS[SKIPPED])
        status_cell.font = STATUS_FONTS.get(outcome.status, STATUS_FONTS[SKIPPED])
        status_cell.alignment = Alignment(horizontal="center", vertical="top")

        sheet.cell(row, headers.index("Why") + 1).alignment = Alignment(
            wrap_text=True, vertical="top")

        for header in ("In toolbar", "In model builder",
                       "In command palette", "In installed CDO"):
            cell = sheet.cell(row, headers.index(header) + 1)
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.font = {"yes": YES_FONT, "NO": NO_FONT}.get(cell.value, UNKNOWN_FONT)

    _finish_table(sheet, len(outcomes), len(headers))


def _issues_sheet(sheet: Worksheet, report: RunReport) -> None:
    headers = ["#", "Issue type", "Operator", "Why", "Category", "Signature",
               "Parameters used", "Exit code", "Command"]
    _header_row(sheet, headers, [5, 30, 22, 70, 20, 10, 26, 10, 90])

    if not report.failures:
        cell = sheet.cell(2, 2, "No failures — every operator that ran, passed.")
        cell.font = STATUS_FONTS[PASS]
        _finish_table(sheet, 0, len(headers))
        return

    # Grouped by root cause, biggest group first: the sheet should open on the
    # thing worth fixing, not on whatever sorts first alphabetically.
    order: dict = {}
    for outcome in report.failures:
        order.setdefault(outcome.issue_type or "Unclassified", []).append(outcome)

    row = 2
    index = 1
    for issue_type, group in sorted(order.items(), key=lambda item: -len(item[1])):
        for outcome in sorted(group, key=lambda item: item.operator):
            values = [index, issue_type, outcome.operator, outcome.why,
                      outcome.category, outcome.signature, outcome.parameters,
                      outcome.returncode, outcome.command]
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row, column_index, _safe(value))
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="top",
                                           wrap_text=column_index == 4)
            row += 1
            index += 1

    _finish_table(sheet, row - 2, len(headers))


def _surfaces_sheet(sheet: Worksheet, report: RunReport) -> None:
    headers = ["#", "Operator", "Category", "In toolbar", "Toolbar placement",
               "In model builder", "In command palette", "In installed CDO",
               "In static schema"]
    _header_row(sheet, headers, [5, 22, 20, 11, 18, 15, 17, 15, 15])

    missing = report.unreachable
    if not missing:
        cell = sheet.cell(2, 2, "Every operator is reachable from the toolbar "
                               "menus, the command palette and the model builder.")
        cell.font = STATUS_FONTS[PASS]
        _finish_table(sheet, 0, len(headers))
        return

    for index, outcome in enumerate(missing, start=1):
        surfaces = outcome.surfaces
        values = [index, outcome.operator, outcome.category,
                  _tick(surfaces.toolbar), surfaces.toolbar_placement,
                  _tick(surfaces.model_builder), _tick(surfaces.palette),
                  _tick(surfaces.installed), "yes" if surfaces.in_schema else "NO"]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(index + 1, column_index, _safe(value))
            cell.border = CELL_BORDER
            if value in ("yes", "NO", "?"):
                cell.alignment = Alignment(horizontal="center")
                cell.font = {"yes": YES_FONT, "NO": NO_FONT}.get(value, UNKNOWN_FONT)

    _finish_table(sheet, len(missing), len(headers))


def _environment_sheet(sheet: Worksheet, report: RunReport) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 100

    row = _section(sheet, 1, "Run")
    for label, value in (
        ("Started", f"{report.started:%Y-%m-%d %H:%M:%S}"),
        ("Finished", f"{report.finished:%Y-%m-%d %H:%M:%S}"),
        ("Wall clock", f"{report.duration:.1f}s"),
        ("CDO binary", report.cdo_binary),
        ("CDO version", report.cdo_version),
        ("Sample inputs", report.sample_description),
        ("Output directory", str(report.output_dir)),
        # What a re-run will cost in disk, in the workbook rather than
        # discovered on a full one.
        ("Output written", _megabytes(report.bytes_written)),
        ("Output pruned afterwards",
         f"{_megabytes(report.bytes_pruned)} — the largest, to fit the "
         f"{OUTPUT_BUDGET_BYTES // (1024 * 1024)} MB budget"
         if report.bytes_pruned else "none"),
        ("Output kept", _megabytes(report.bytes_written - report.bytes_pruned)),
    ):
        sheet.cell(row, 1, label).font = Font(bold=True)
        sheet.cell(row, 2, value).alignment = Alignment(wrap_text=True)
        row += 1

    if report.surface_errors:
        row = _section(sheet, row + 1, "Surface inspection problems")
        for surface, message in report.surface_errors.items():
            sheet.cell(row, 1, surface).font = Font(bold=True)
            sheet.cell(row, 2, _safe(message)).alignment = Alignment(wrap_text=True)
            row += 1

    row = _section(sheet, row + 1, "Reading this report")
    for label, value in (
        ("Status", "pass = CDO ran it and the integration judged the result good. "
                   "fail = it ran and did not. skipped = it cannot be driven from "
                   "files alone; the reason is in the Why column."),
        ("Input ext / Output ext",
         "The extension each operator's files usually carry. An informational "
         "operator writes to stdout, so its output is captured as .txt."),
        ("Surface columns",
         "yes / NO / ? — '?' means the surface could not be inspected in this "
         "run, which is not the same as the operator being absent."),
    ):
        sheet.cell(row, 1, label).font = Font(bold=True)
        sheet.cell(row, 2, value).alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[row].height = 45
        row += 1


# --------------------------------------------------------------------------
# Shared formatting
# --------------------------------------------------------------------------

def _header_row(sheet: Worksheet, headers: Sequence[str],
                widths: Sequence[int]) -> None:
    for index, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(1, index, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 30


def _finish_table(sheet: Worksheet, rows: int, columns: int) -> None:
    """Freeze the header and put a filter on it.

    Both matter on a 943-row sheet: without the freeze the header scrolls away
    in the first flick, and without the filter the reader has to sort by hand
    to answer "show me the failures".
    """
    sheet.freeze_panes = "B2"
    last = get_column_letter(columns)
    sheet.auto_filter.ref = f"A1:{last}{max(rows, 1) + 1}"
